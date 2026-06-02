from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

from indexops.ocr import extract_pdf_text_with_ocr, extract_text_with_ocr

SUPPORTED_EXTENSIONS = {
    ".pdf",
    ".png",
    ".jpg",
    ".jpeg",
    ".tif",
    ".tiff",
    ".txt",
    ".csv",
    ".xlsx",
    ".xls",
    ".docx",
}
TEXT_EXTENSIONS = {".txt", ".csv", ".docx", ".xlsx", ".xls", ".pdf"}
OCR_CACHE_EXTENSIONS = {".pdf", ".png", ".jpg", ".jpeg", ".tif", ".tiff"}


@dataclass(frozen=True)
class DocumentText:
    text: str
    status: str
    metadata: dict[str, str] = field(default_factory=dict)


def read_document_text(path: Path, *, inline_ocr: bool = False, max_ocr_pages: int = 10) -> DocumentText:
    suffix = path.suffix.lower()
    if suffix not in SUPPORTED_EXTENSIONS:
        return DocumentText(text="", status="EXTENSION_NO_SOPORTADA")

    try:
        if suffix == ".pdf":
            return _read_pdf(path, inline_ocr=inline_ocr, max_ocr_pages=max_ocr_pages)
        if suffix in {".png", ".jpg", ".jpeg", ".tif", ".tiff"}:
            if not inline_ocr:
                return DocumentText(text="", status="IMAGEN_REQUIERE_OCR")
            ocr = extract_text_with_ocr(path)
            return DocumentText(text=ocr.text, status=ocr.status)
        if suffix in {".txt", ".csv"}:
            return DocumentText(text=path.read_text(encoding="utf-8", errors="ignore"), status="TEXTO_OK")
        if suffix in {".xlsx", ".xls"}:
            return _read_excel(path)
        if suffix == ".docx":
            return _read_docx(path)
    except Exception as exc:
        return DocumentText(text="", status=f"LECTURA_ERROR: {exc}")

    return DocumentText(text="", status="LECTURA_NO_IMPLEMENTADA")


def _read_pdf(path: Path, *, inline_ocr: bool, max_ocr_pages: int) -> DocumentText:
    try:
        from pypdf import PdfReader  # type: ignore
    except ImportError:
        return DocumentText(text="", status="PDF_EXTRACCION_NO_DISPONIBLE")

    try:
        reader = PdfReader(str(path))
        if reader.is_encrypted:
            return DocumentText(text="", status="PDF_PROTEGIDO")
        pages = [(page.extract_text() or "") for page in reader.pages]
    except Exception as exc:
        return DocumentText(text="", status=f"PDF_ERROR: {exc}")

    text = "\n".join(pages).strip()
    if text:
        return DocumentText(text=text, status=f"PDF_TEXTO_OK;paginas={len(pages)}")
    if inline_ocr:
        ocr = extract_pdf_text_with_ocr(path, max_pages=max_ocr_pages)
        return DocumentText(text=ocr.text, status=ocr.status)
    return DocumentText(text="", status="PDF_ESCANEADO_PROBABLE")


def _read_docx(path: Path) -> DocumentText:
    from docx import Document  # type: ignore

    doc = Document(str(path))
    text = "\n".join(p.text for p in doc.paragraphs if p.text).strip()
    return DocumentText(text=text, status="DOCX_OK" if text else "DOCX_SIN_TEXTO")


def _read_excel(path: Path) -> DocumentText:
    from openpyxl import load_workbook  # type: ignore

    wb = load_workbook(str(path), read_only=True, data_only=True)
    chunks: list[str] = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(max_row=200, values_only=True):
            for cell in row:
                if cell is not None:
                    chunks.append(str(cell))
    wb.close()
    text = " ".join(chunks).strip()
    return DocumentText(text=text, status="XLSX_OK" if text else "XLSX_SIN_TEXTO")


def preview_text(text: str, limit: int = 400) -> str:
    rendered = " ".join(text.split())
    if len(rendered) <= limit:
        return rendered
    return rendered[: limit - 3] + "..."
