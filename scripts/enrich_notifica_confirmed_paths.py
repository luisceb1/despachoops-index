from __future__ import annotations

import argparse
import csv
import json
import re
import unicodedata
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_CONFIRMED_CSV = r"\\Luiscp\d\DespachoOpsData\Index\notifica_assignments_confirmed.csv"
DEFAULT_OUT_XLSX = r"\\Luiscp\d\DespachoOpsData\Index\notifica_assignments_confirmed.xlsx"
DEFAULT_CLIENTES_ROOT = r"\\Luiscp\d\Cebrian y Fraile Abogados\Clientes"

DIAGNOSTIC_FIELDS = [
    "cliente_folder_detected",
    "cliente_folder_match_score",
    "cliente_folder_match_reason",
]

GENERIC_NAME_TOKENS = {
    "sl",
    "sociedad",
    "limitada",
    "de",
    "del",
    "la",
    "el",
    "los",
    "las",
    "y",
}

AUTO_EXISTING_NOTE = "Destino Notificaciones existente propuesto automaticamente."
AUTO_MISSING_NOTE = "Carpeta Notificaciones no existe; destino propuesto pendiente de crear/confirmar."
NO_CLIENT_FOLDER_NOTE = "No se pudo localizar carpeta cliente en Clientes root."
CONFLICT_NOTE = "Conflicto de carpeta cliente."


class FolderMatch:
    def __init__(self, folder: Path | None, score: int = 0, reason: str = "", conflict: bool = False):
        self.folder = folder
        self.score = score
        self.reason = reason
        self.conflict = conflict


def strip_sep_line(lines: list[str]) -> list[str]:
    if lines and lines[0].strip().lower() == "sep=;":
        return lines[1:]
    return lines


def detect_delimiter(lines: list[str]) -> str:
    data_lines = strip_sep_line(lines)
    sample = "\n".join(data_lines[:20])

    if sample:
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
            if dialect.delimiter in {";", ",", "\t"}:
                return dialect.delimiter
        except csv.Error:
            pass

    header = data_lines[0] if data_lines else ""
    comma_count = header.count(",")
    semicolon_count = header.count(";")
    tab_count = header.count("\t")

    if tab_count > comma_count and tab_count > semicolon_count:
        return "\t"
    if comma_count > semicolon_count:
        return ","
    if semicolon_count > comma_count:
        return ";"
    return ";"


def read_csv_flexible(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    last_error: Exception | None = None

    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin1"):
        try:
            text = path.read_text(encoding=encoding)
            break
        except UnicodeDecodeError as exc:
            last_error = exc
    else:
        raise last_error or UnicodeDecodeError("utf-8", b"", 0, 1, "unknown error")

    raw_lines = text.splitlines()
    lines = strip_sep_line(raw_lines)
    delimiter = detect_delimiter(raw_lines)
    reader = csv.DictReader(lines, delimiter=delimiter)
    fieldnames = [field.strip() for field in (reader.fieldnames or [])]
    rows = [
        {(key or "").strip(): (value or "").strip() for key, value in row.items()}
        for row in reader
    ]
    return fieldnames, rows


def ensure_fieldnames(fieldnames: list[str]) -> list[str]:
    output = list(fieldnames)
    for field in DIAGNOSTIC_FIELDS:
        if field not in output:
            output.append(field)
    return output


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)

    with path.open("w", newline="", encoding="utf-8-sig") as f:
        f.write("sep=;\n")
        writer = csv.DictWriter(
            f,
            fieldnames=fieldnames,
            delimiter=";",
            lineterminator="\n",
        )
        writer.writeheader()
        writer.writerows({field: row.get(field, "") for field in fieldnames} for row in rows)


def write_xlsx(path: Path, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Confirmaciones"
    sheet.append(fieldnames)

    for row in rows:
        sheet.append([row.get(field, "") for field in fieldnames])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    sheet.freeze_panes = "A2"
    last_column = get_column_letter(len(fieldnames))
    last_row = max(sheet.max_row, 1)
    sheet.auto_filter.ref = f"A1:{last_column}{last_row}"

    for column_idx, field in enumerate(fieldnames, start=1):
        values = [field, *(str(row.get(field, "")) for row in rows[:200])]
        width = min(max(len(value) for value in values) + 2, 70)
        sheet.column_dimensions[get_column_letter(column_idx)].width = width

    workbook.save(path)


def normalize_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", (value or "").strip().lower()).strip("_")


def normalize_name(value: str) -> str:
    value = unicodedata.normalize("NFKD", value or "")
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    value = value.upper()
    value = re.sub(r"\bS\s*\.?\s*L\s*\.?\b", " SL ", value)
    value = re.sub(r"\bSOCIEDAD\s+LIMITADA\b", " SL ", value)
    value = value.lower()
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return " ".join(value.split())


def significant_tokens(value: str) -> set[str]:
    return {
        token
        for token in normalize_name(value).split()
        if token not in GENERIC_NAME_TOKENS and len(token) >= 2
    }


def compact(value: str) -> str:
    return normalize_name(value).replace(" ", "")


def walk_json_values(value: object) -> Iterable[object]:
    yield value
    if isinstance(value, dict):
        for child in value.values():
            yield from walk_json_values(child)
    elif isinstance(value, list):
        for child in value:
            yield from walk_json_values(child)


def load_context_names(json_path: str) -> tuple[str, str]:
    if not json_path:
        return "", ""

    try:
        data = json.loads(Path(json_path).read_text(encoding="utf-8", errors="replace"))
    except Exception:
        return "", ""

    if not isinstance(data, dict):
        return "", ""

    client_name = str(data.get("client_name") or "").strip()
    client_key = str(data.get("client_key") or "").strip()
    return client_name, client_key


def folder_candidates_from_row(row: dict[str, str]) -> tuple[list[str], str, str]:
    client_name, client_key = load_context_names(row.get("client_context_json", ""))
    candidates = [
        row.get("cliente_sugerido", ""),
        client_name,
        client_key.replace("_", " "),
    ]

    cleaned = []
    seen = set()
    for candidate in candidates:
        candidate = candidate.strip()
        key = normalize_name(candidate)
        if not candidate or key in seen:
            continue
        seen.add(key)
        cleaned.append(candidate)

    return cleaned, client_name, client_key


def scan_cliente_folders(clientes_root: Path) -> list[Path]:
    if not clientes_root.exists():
        return []
    return sorted(path for path in clientes_root.iterdir() if path.is_dir())


def score_folder(candidate: str, client_key: str, folder: Path) -> tuple[int, str]:
    candidate_norm = normalize_name(candidate)
    folder_norm = normalize_name(folder.name)
    if not candidate_norm or not folder_norm:
        return 0, ""

    if candidate_norm == folder_norm:
        return 100, f"nombre exacto: {candidate}"

    key_norm = compact(client_key.replace("_", " "))
    if key_norm and key_norm == compact(folder.name):
        return 85, f"client_key coincide: {client_key}"

    candidate_tokens = significant_tokens(candidate)
    folder_tokens = significant_tokens(folder.name)
    if not candidate_tokens or not folder_tokens:
        return 0, ""

    hits = sorted(candidate_tokens & folder_tokens)
    if candidate_tokens <= folder_tokens:
        return 70, "todos los tokens: " + ", ".join(sorted(candidate_tokens))
    if len(hits) >= 2:
        return 50, "tokens parciales: " + ", ".join(hits)

    return 0, ""


def match_cliente_folder(row: dict[str, str], folders: list[Path]) -> FolderMatch:
    candidates, _client_name, client_key = folder_candidates_from_row(row)
    scored: list[tuple[int, Path, str]] = []

    for candidate in candidates:
        for folder in folders:
            score, reason = score_folder(candidate, client_key, folder)
            if score:
                scored.append((score, folder, reason))

    if not scored:
        return FolderMatch(folder=None)

    scored.sort(key=lambda item: item[0], reverse=True)
    best_score = scored[0][0]
    best = [item for item in scored if item[0] == best_score]
    unique_best = {str(item[1]).lower() for item in best}
    if len(unique_best) > 1:
        return FolderMatch(folder=None, score=best_score, reason=CONFLICT_NOTE, conflict=True)

    _score, folder, reason = best[0]
    return FolderMatch(folder=folder, score=best_score, reason=reason)


def append_observation(existing: str, note: str) -> str:
    existing = (existing or "").strip()
    if note in existing:
        return existing
    if not existing:
        return note
    return existing + " | " + note


def should_process(row: dict[str, str], force: bool) -> bool:
    if (row.get("confirmado") or "").strip().upper() == "SI":
        return False
    if row.get("confirmed_expediente_path") and not force:
        return False
    return True


def enrich_rows(rows: list[dict[str, str]], clientes_root: Path, force: bool) -> dict[str, int]:
    counts = {
        "folders": 0,
        "existing": 0,
        "missing": 0,
        "none": 0,
        "human": 0,
    }
    folders = scan_cliente_folders(clientes_root)

    for row in rows:
        row.setdefault("cliente_folder_detected", "")
        row.setdefault("cliente_folder_match_score", "")
        row.setdefault("cliente_folder_match_reason", "")

        if (row.get("confirmado") or "").strip().upper() == "SI":
            counts["human"] += 1
            continue

        if row.get("confirmed_expediente_path") and not force:
            counts["human"] += 1
            continue

        if not should_process(row, force):
            continue

        match = match_cliente_folder(row, folders)
        row["cliente_folder_match_score"] = str(match.score) if match.score else ""
        row["cliente_folder_match_reason"] = match.reason

        if match.conflict:
            row["accion"] = "elegir_expediente"
            row["observaciones"] = append_observation(row.get("observaciones", ""), CONFLICT_NOTE)
            counts["none"] += 1
            continue

        if not match.folder or match.score < 70:
            row["accion"] = "elegir_expediente"
            row["observaciones"] = append_observation(row.get("observaciones", ""), NO_CLIENT_FOLDER_NOTE)
            counts["none"] += 1
            continue

        counts["folders"] += 1
        row["cliente_folder_detected"] = str(match.folder)
        target = match.folder / "Notificaciones"
        row["confirmed_expediente_path"] = str(target)
        row["confirmado"] = "NO"

        if target.exists():
            row["accion"] = "confirmar_archivo"
            row["observaciones"] = append_observation(row.get("observaciones", ""), AUTO_EXISTING_NOTE)
            counts["existing"] += 1
        else:
            row["accion"] = "crear_carpeta_y_confirmar"
            row["observaciones"] = append_observation(row.get("observaciones", ""), AUTO_MISSING_NOTE)
            counts["missing"] += 1

    return counts


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Propone confirmed_expediente_path en la plantilla de confirmaciones Notifica."
    )
    parser.add_argument("--confirmed-csv", default=DEFAULT_CONFIRMED_CSV)
    parser.add_argument("--clientes-root", default=DEFAULT_CLIENTES_ROOT)
    parser.add_argument("--out-xlsx", default=DEFAULT_OUT_XLSX)
    parser.add_argument("--force", action="store_true")
    args = parser.parse_args()

    confirmed_csv = Path(args.confirmed_csv)
    clientes_root = Path(args.clientes_root)
    out_xlsx = Path(args.out_xlsx)

    fieldnames, rows = read_csv_flexible(confirmed_csv)
    fieldnames = ensure_fieldnames(fieldnames)
    counts = enrich_rows(rows, clientes_root=clientes_root, force=args.force)

    write_csv(confirmed_csv, fieldnames, rows)
    write_xlsx(out_xlsx, fieldnames, rows)

    print("Filas leidas:", len(rows))
    print("Carpetas cliente detectadas:", counts["folders"])
    print("Destinos Notificaciones existentes:", counts["existing"])
    print("Destinos Notificaciones no existentes:", counts["missing"])
    print("Sin carpeta cliente:", counts["none"])
    print("Confirmaciones humanas conservadas:", counts["human"])
    print("CSV:", confirmed_csv)
    print("XLSX:", out_xlsx)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
