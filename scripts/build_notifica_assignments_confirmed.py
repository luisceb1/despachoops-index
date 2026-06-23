from __future__ import annotations

import argparse
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_CANDIDATES = r"\\Luiscp\d\DespachoOpsData\Index\notifica_assignment_candidates.csv"
DEFAULT_OUT_CSV = r"\\Luiscp\d\DespachoOpsData\Index\notifica_assignments_confirmed.csv"
DEFAULT_OUT_XLSX = r"\\Luiscp\d\DespachoOpsData\Index\notifica_assignments_confirmed.xlsx"

OUTPUT_FIELDS = [
    "notification_id",
    "source",
    "received_at",
    "assignment_nif",
    "assignment_nif_source",
    "entity_nif",
    "certificate_nif",
    "represented_nif",
    "cliente_sugerido",
    "client_context_json",
    "client_match_score",
    "client_match_reason",
    "estado_asignacion",
    "expediente_sugerido",
    "expediente_path",
    "title",
    "confirmed_cliente",
    "confirmed_expediente_path",
    "accion",
    "confirmado",
    "revisado_por",
    "fecha_revision",
    "observaciones",
]

HUMAN_FIELDS = {
    "confirmed_cliente",
    "confirmed_expediente_path",
    "accion",
    "confirmado",
    "revisado_por",
    "fecha_revision",
    "observaciones",
}


def strip_sep_line(lines: list[str]) -> list[str]:
    if lines and lines[0].strip().lower() == "sep=;":
        return lines[1:]
    return lines


def detect_delimiter(lines: list[str]) -> str:
    data_lines = strip_sep_line(lines)
    sample = "\n".join(data_lines[:20])

    if sample:
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
            if dialect.delimiter in {";", ",", "\t"}:
                return dialect.delimiter
        except csv.Error:
            pass

    header = data_lines[0] if data_lines else ""
    comma_count = header.count(",")
    semicolon_count = header.count(";")
    tab_count = header.count("\t")

    if tab_count > comma_count and tab_count > semicolon_count:
        return "\t"
    if comma_count > semicolon_count:
        return ","
    if semicolon_count > comma_count:
        return ";"
    return ";"


def read_csv_flexible(path: Path) -> list[dict[str, str]]:
    last_error: Exception | None = None

    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin1"):
        try:
            text = path.read_text(encoding=encoding)
            break
        except UnicodeDecodeError as exc:
            last_error = exc
    else:
        raise last_error or UnicodeDecodeError("utf-8", b"", 0, 1, "unknown error")

    raw_lines = text.splitlines()
    lines = strip_sep_line(raw_lines)
    delimiter = detect_delimiter(raw_lines)
    reader = csv.DictReader(lines, delimiter=delimiter)
    return [
        {(key or "").strip(): (value or "").strip() for key, value in row.items()}
        for row in reader
    ]


def action_for_estado(estado: str) -> str:
    estado = (estado or "").strip()
    if estado == "pendiente_confirmacion":
        return "confirmar_archivo"
    if estado == "cliente_identificado":
        return "elegir_expediente"
    if estado == "conflicto_identidad":
        return "resolver_conflicto"
    return "revisar_manualmente"


def build_base_row(candidate: dict[str, str]) -> dict[str, str]:
    row = {field: candidate.get(field, "") for field in OUTPUT_FIELDS}
    row["confirmed_cliente"] = candidate.get("cliente_sugerido", "")
    row["confirmed_expediente_path"] = candidate.get("expediente_path", "")
    row["accion"] = action_for_estado(candidate.get("estado_asignacion", ""))
    row["confirmado"] = "NO"
    row["revisado_por"] = ""
    row["fecha_revision"] = ""
    row["observaciones"] = ""
    return row


def merge_existing_human_fields(
    base_row: dict[str, str],
    existing_by_id: dict[str, dict[str, str]],
) -> tuple[dict[str, str], bool]:
    notification_id = base_row.get("notification_id", "")
    existing = existing_by_id.get(notification_id)
    if not existing:
        return base_row, False

    merged = dict(base_row)
    for field in HUMAN_FIELDS:
        merged[field] = existing.get(field, merged.get(field, ""))
    return merged, True


def build_confirmed_rows(
    candidate_rows: list[dict[str, str]],
    existing_rows: list[dict[str, str]],
) -> tuple[list[dict[str, str]], int, int]:
    existing_by_id = {
        row.get("notification_id", ""): row
        for row in existing_rows
        if row.get("notification_id", "")
    }

    output_rows = []
    new_count = 0
    conserved_count = 0

    for candidate in candidate_rows:
        base_row = build_base_row(candidate)
        merged, conserved = merge_existing_human_fields(base_row, existing_by_id)
        output_rows.append({field: merged.get(field, "") for field in OUTPUT_FIELDS})
        if conserved:
            conserved_count += 1
        else:
            new_count += 1

    return output_rows, new_count, conserved_count


def write_csv(rows: list[dict[str, str]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with output_path.open("w", newline="", encoding="utf-8-sig") as f:
        f.write("sep=;\n")
        writer = csv.DictWriter(
            f,
            fieldnames=OUTPUT_FIELDS,
            delimiter=";",
            lineterminator="\n",
        )
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(rows: list[dict[str, str]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Confirmaciones"
    sheet.append(OUTPUT_FIELDS)

    for row in rows:
        sheet.append([row.get(field, "") for field in OUTPUT_FIELDS])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    sheet.freeze_panes = "A2"
    last_column = get_column_letter(len(OUTPUT_FIELDS))
    last_row = max(sheet.max_row, 1)
    sheet.auto_filter.ref = f"A1:{last_column}{last_row}"

    for column_idx, field in enumerate(OUTPUT_FIELDS, start=1):
        values = [field, *(str(row.get(field, "")) for row in rows[:200])]
        width = min(max(len(value) for value in values) + 2, 70)
        sheet.column_dimensions[get_column_letter(column_idx)].width = width

    workbook.save(output_path)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Crea o actualiza la plantilla humana de confirmacion de asignaciones Notifica."
    )
    parser.add_argument("--candidates", default=DEFAULT_CANDIDATES)
    parser.add_argument("--out-csv", default=DEFAULT_OUT_CSV)
    parser.add_argument("--out-xlsx", default=DEFAULT_OUT_XLSX)
    args = parser.parse_args()

    candidates_path = Path(args.candidates)
    out_csv = Path(args.out_csv)
    out_xlsx = Path(args.out_xlsx)

    candidate_rows = read_csv_flexible(candidates_path)
    existing_rows = read_csv_flexible(out_csv) if out_csv.exists() else []
    confirmed_rows, new_count, conserved_count = build_confirmed_rows(candidate_rows, existing_rows)

    write_csv(confirmed_rows, out_csv)
    write_xlsx(confirmed_rows, out_xlsx)

    print("Filas candidates:", len(candidate_rows))
    print("Filas confirmed final:", len(confirmed_rows))
    print("Nuevas anadidas:", new_count)
    print("Existentes conservadas:", conserved_count)
    print("CSV:", out_csv)
    print("XLSX:", out_xlsx)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
