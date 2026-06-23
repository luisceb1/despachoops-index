from __future__ import annotations

import argparse
import csv
import json
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_NOTIFICA_INBOX = r"C:\DespachoOpsData\Notifica\notificaciones_inbox.csv"
DEFAULT_LIVE_EXPEDIENTES = r"D:\DespachoOpsData\Index\live_expedientes_index.csv"
DEFAULT_CLIENT_CONTEXT_DIR = r"D:\DespachoOpsData\Index\client_context_index"
DEFAULT_OUT_CSV = r"D:\DespachoOpsData\Index\notifica_assignment_candidates.csv"
DEFAULT_OUT_XLSX = r"D:\DespachoOpsData\Index\notifica_assignment_candidates.xlsx"

OUTPUT_FIELDS = [
    "notification_id",
    "source",
    "received_at",
    "entity_nif",
    "entity_name",
    "title",
    "download_status",
    "main_pdf_path",
    "receipt_path",
    "client_context_json",
    "client_match_score",
    "client_match_reason",
    "cliente_sugerido",
    "expediente_sugerido",
    "expediente_path",
    "match_score",
    "match_reason",
    "estado_asignacion",
    "accion_sugerida",
    "observaciones",
    "index_columns_detected",
    "match_debug",
]

STOP_TOKENS = {
    "2020",
    "2021",
    "2022",
    "2023",
    "2024",
    "2025",
    "2026",
    "2027",
    "2028",
    "2029",
    "2030",
    "para",
    "por",
    "con",
    "sin",
    "del",
    "las",
    "los",
    "una",
    "uno",
    "unos",
    "unas",
    "este",
    "esta",
    "estos",
    "estas",
    "notificacion",
    "comunicacion",
    "procedimiento",
    "expediente",
    "administracion",
    "administrativa",
    "electronica",
    "aceptar",
    "rechazar",
    "fecha",
    "caducidad",
    "si",
    "no",
    "pendiente",
    "comparecencia",
    "agencia",
    "estatal",
    "tesoreria",
    "general",
    "seguridad",
    "social",
    "aeat",
    "tributaria",
    "notificar",
    "notificaciones",
    "genericas",
    "administrati",
    "ayuntamiento",
    "organismo",
    "postal",
    "pago",
    "pagos",
    "devoluciones",
}

COMPANY_FORM_TOKENS = {
    "sl",
    "sll",
    "slu",
    "sa",
    "sau",
    "sociedad",
    "limitada",
    "anonima",
}

REPRESENTATIVE_TOKENS = {
    "juan",
    "carlos",
    "garcia",
}

INDEX_COLUMN_ALIASES = {
    "cliente": (
        "cliente",
        "client",
        "nombre_cliente",
        "cliente_nombre",
        "display_name",
        "nombre",
        "cliente_estimado",
        "cliente_carpeta",
    ),
    "expediente": (
        "expediente",
        "expediente_nombre",
        "nombre_expediente",
        "asunto",
        "contains",
    ),
    "ruta": (
        "path",
        "expediente_path",
        "ruta",
        "folder",
        "folder_path",
        "expediente_carpeta",
        "md_path",
    ),
    "nif": (
        "nif",
        "dni",
        "cif",
        "nie",
        "tax_id",
        "cliente_nif",
    ),
}

NOTIFICATION_ALIASES = {
    "notification_id": ("notification_id", "id", "notificacion_id"),
    "source": ("source", "origen", "buzon"),
    "received_at": ("received_at", "fecha_recepcion", "received"),
    "entity_nif": ("entity_nif", "nif", "cif", "dni", "destinatario_nif"),
    "entity_name": (
        "entity_name",
        "entity",
        "cliente",
        "razon_social",
        "destinatario",
    ),
    "title": ("title", "titulo", "asunto", "subject"),
    "download_status": ("download_status", "estado_descarga", "status"),
    "main_pdf_path": ("main_pdf_path", "pdf_path", "document_path", "documento_path"),
    "receipt_path": ("receipt_path", "acuse_path", "justificante_path"),
}

CLIENT_NAME_KEYS = {
    "client_name",
    "cliente",
    "nombre_cliente",
    "cliente_nombre",
    "display_name",
    "name",
    "nombre",
    "razon_social",
}


@dataclass(frozen=True)
class Match:
    row: dict[str, str] | None
    score: int
    reasons: list[str]
    debug: str = ""


@dataclass(frozen=True)
class ClientContext:
    json_path: Path
    stem: str
    client_name: str
    normalized_text: str
    normalized_nif_text: str
    text_tokens: set[str]
    path_tokens: set[str]
    name_tokens: set[str]


@dataclass(frozen=True)
class ClientMatch:
    context: ClientContext | None
    score: int
    reasons: list[str]
    debug: str = ""


@dataclass(frozen=True)
class IndexColumns:
    cliente: list[str]
    expediente: list[str]
    ruta: list[str]
    nif: list[str]

    def describe(self) -> str:
        parts = []
        for name in ("cliente", "expediente", "ruta", "nif"):
            values = getattr(self, name)
            parts.append(f"{name}={','.join(values) if values else '-'}")
        return " | ".join(parts)


@dataclass(frozen=True)
class CsvReadResult:
    rows: list[dict[str, str]]
    fieldnames: list[str]
    delimiter: str


def fix_mojibake(value: str) -> str:
    if not isinstance(value, str):
        return value

    if "\u00c3" in value or "\u00c2" in value:
        try:
            return value.encode("latin1").decode("utf-8")
        except Exception:
            return value

    return value


def clean_row(row: dict[str, str]) -> dict[str, str]:
    return {
        (key or "").strip(): fix_mojibake(value or "").strip()
        for key, value in row.items()
    }


def strip_sep_line(lines: list[str]) -> list[str]:
    if lines and lines[0].strip().lower() == "sep=;":
        return lines[1:]
    return lines


def detect_delimiter(lines: list[str]) -> str:
    data_lines = strip_sep_line(lines)
    sample = "\n".join(data_lines[:20])

    if sample:
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
            if dialect.delimiter in {";", ",", "\t"}:
                return dialect.delimiter
        except csv.Error:
            pass

    header = data_lines[0] if data_lines else ""
    comma_count = header.count(",")
    semicolon_count = header.count(";")
    tab_count = header.count("\t")

    if tab_count > comma_count and tab_count > semicolon_count:
        return "\t"
    if comma_count > semicolon_count:
        return ","
    if semicolon_count > comma_count:
        return ";"
    return ";"


def read_csv_flexible(path: Path) -> CsvReadResult:
    last_error: Exception | None = None

    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin1"):
        try:
            text = path.read_text(encoding=encoding)
            break
        except UnicodeDecodeError as exc:
            last_error = exc
    else:
        raise last_error or UnicodeDecodeError("utf-8", b"", 0, 1, "unknown error")

    lines = strip_sep_line(text.splitlines())
    delimiter = detect_delimiter(text.splitlines())
    reader = csv.DictReader(lines, delimiter=delimiter)
    rows = [clean_row(row) for row in reader]
    fieldnames = [field.strip() for field in (reader.fieldnames or [])]
    return CsvReadResult(rows=rows, fieldnames=fieldnames, delimiter=delimiter)


def normalize(value: str) -> str:
    value = fix_mojibake(value or "")
    value = unicodedata.normalize("NFKD", value)
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    value = value.upper()
    value = re.sub(r"\bS\s*\.?\s*L\s*\.?\b", " SL ", value)
    value = re.sub(r"\bS\s*\.?\s*A\s*\.?\b", " SA ", value)
    value = value.lower()
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return " ".join(value.split())


def normalize_nif(value: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", (value or "").upper())


def extract_nifs(value: str) -> set[str]:
    value = normalize(value).upper()
    patterns = (
        r"\b[XYZ]\d{7}[A-Z]\b",
        r"\b\d{8}[A-Z]\b",
        r"\b[A-Z]\d{7,8}[A-Z0-9]\b",
    )
    found: set[str] = set()
    for pattern in patterns:
        found.update(re.findall(pattern, value))
    return {normalize_nif(item) for item in found if normalize_nif(item)}


def remove_nifs(value: str) -> str:
    value = re.sub(r"\b[XYZ]?\d{7,8}[A-Z]\b", " ", value.upper())
    value = re.sub(r"\b[A-Z]\d{7,8}[A-Z0-9]\b", " ", value)
    return value


def is_noise_token(token: str, allow_short_nif: bool = False) -> bool:
    if token in STOP_TOKENS:
        return True
    if token in COMPANY_FORM_TOKENS:
        return True
    if re.fullmatch(r"20[2-3][0-9]", token):
        return True
    if token.isdigit():
        return True
    if re.fullmatch(r"[a-z]\d{7,8}[a-z0-9]", token):
        return True
    if re.fullmatch(r"[xyz]?\d{7,8}[a-z]", token):
        return True
    if len(token) < 4 and not allow_short_nif:
        return True
    return False


def significant_tokens(*values: str, extra_stop: set[str] | None = None) -> set[str]:
    tokens: set[str] = set()
    extra_stop = extra_stop or set()

    for value in values:
        for token in normalize(value).split():
            if token in extra_stop:
                continue
            if is_noise_token(token):
                continue
            tokens.add(token)

    return tokens


def name_tokens(value: str, extra_stop: set[str] | None = None) -> list[str]:
    normalized = normalize(remove_nifs(value))
    extra_stop = extra_stop or set()
    return [
        token
        for token in normalized.split()
        if token not in extra_stop and not is_noise_token(token)
    ]


def name_variants(value: str, extra_stop: set[str] | None = None) -> set[str]:
    tokens = name_tokens(value, extra_stop=extra_stop)
    if not tokens:
        return set()
    variants = {" ".join(tokens), " ".join(sorted(tokens))}
    return {variant for variant in variants if variant}


def display_name_from_stem(stem: str) -> str:
    return " ".join(part.capitalize() for part in stem.replace("_", " ").split())


def first_present(row: dict[str, str], aliases: Iterable[str]) -> str:
    normalized_keys = {normalize(key): key for key in row}

    for alias in aliases:
        key = normalized_keys.get(normalize(alias))
        if key and row.get(key):
            return row[key].strip()

    return ""


def notification_value(row: dict[str, str], canonical: str) -> str:
    return first_present(row, NOTIFICATION_ALIASES[canonical])


def representative_nif(entity_name: str) -> str:
    match = re.search(r"\(?\s*R\s*:\s*([A-Z]\d{7,8}[A-Z0-9])", entity_name or "", flags=re.I)
    return normalize_nif(match.group(1)) if match else ""


def notification_client_nif(row: dict[str, str]) -> str:
    entity_name = notification_value(row, "entity_name")
    return representative_nif(entity_name) or normalize_nif(notification_value(row, "entity_nif"))


def notification_client_values(row: dict[str, str]) -> list[str]:
    return [
        notification_value(row, "entity_name"),
        first_present(row, ("cliente_sugerido", "cliente", "cliente_detectado")),
        notification_value(row, "title"),
    ]


def notification_client_tokens(
    row: dict[str, str],
    token_hints: dict[str, set[str]] | None = None,
) -> set[str]:
    entity_name = notification_value(row, "entity_name")
    extra_stop = REPRESENTATIVE_TOKENS if representative_nif(entity_name) else set()
    tokens = significant_tokens(*notification_client_values(row), extra_stop=extra_stop)
    client_nif = notification_client_nif(row)
    if token_hints and client_nif:
        tokens.update(token_hints.get(client_nif, set()))
    return tokens


def build_nif_token_hints(notifications: list[dict[str, str]]) -> dict[str, set[str]]:
    hints: dict[str, set[str]] = {}
    for notification in notifications:
        client_nif = notification_client_nif(notification)
        if not client_nif:
            continue
        hints.setdefault(client_nif, set()).update(notification_client_tokens(notification))
    return hints


def relevant_tokens(*values: str) -> set[str]:
    return significant_tokens(*values)


def detect_index_columns(rows: list[dict[str, str]]) -> IndexColumns:
    if not rows:
        return IndexColumns(cliente=[], expediente=[], ruta=[], nif=[])

    keys = list(rows[0].keys())
    by_norm = {normalize(key): key for key in keys}
    detected: dict[str, list[str]] = {}

    for group, aliases in INDEX_COLUMN_ALIASES.items():
        columns = []
        for alias in aliases:
            key = by_norm.get(normalize(alias))
            if key and key not in columns:
                columns.append(key)
        detected[group] = columns

    return IndexColumns(
        cliente=detected["cliente"],
        expediente=detected["expediente"],
        ruta=detected["ruta"],
        nif=detected["nif"],
    )


def values_from_columns(row: dict[str, str], columns: Iterable[str]) -> list[str]:
    values = []
    for column in columns:
        value = row.get(column, "").strip()
        if value:
            values.append(value)
    return values


def unique_join(values: Iterable[str], separator: str = " | ") -> str:
    cleaned = []
    seen = set()

    for value in values:
        value = (value or "").strip()
        key = normalize(value)
        if not value or key in seen:
            continue
        seen.add(key)
        cleaned.append(value)

    return separator.join(cleaned)


def resolve_existing_path(path: Path) -> Path:
    if path.exists():
        return path

    raw = str(path)
    alternates = []
    if raw.startswith("D:\\DespachoOpsData\\"):
        alternates.append(Path(raw.replace("D:\\DespachoOpsData\\", "\\\\Luiscp\\d\\DespachoOpsData\\", 1)))
    if raw.startswith("\\\\Luiscp\\d\\DespachoOpsData\\"):
        alternates.append(Path(raw.replace("\\\\Luiscp\\d\\DespachoOpsData\\", "D:\\DespachoOpsData\\", 1)))

    for alternate in alternates:
        if alternate.exists():
            return alternate

    return path


def find_json_client_name(data: object) -> str:
    if not isinstance(data, dict):
        return ""

    for key, value in data.items():
        if normalize(str(key)) in CLIENT_NAME_KEYS and isinstance(value, (str, int, float)):
            text = str(value).strip()
            if text:
                return text

    return ""


def load_client_context(path: Path) -> ClientContext:
    raw_text = path.read_text(encoding="utf-8", errors="replace")
    client_name = ""
    try:
        data = json.loads(raw_text)
        client_name = find_json_client_name(data)
        text = json.dumps(data, ensure_ascii=False, sort_keys=True)
    except Exception:
        text = raw_text

    if not client_name:
        client_name = display_name_from_stem(path.stem)

    normalized_text = normalize(text)
    path_text = f"{path.stem} {path.name} {path}"
    return ClientContext(
        json_path=path,
        stem=path.stem,
        client_name=client_name,
        normalized_text=normalized_text,
        normalized_nif_text=normalize_nif(text),
        text_tokens=significant_tokens(text),
        path_tokens=significant_tokens(path_text),
        name_tokens=significant_tokens(client_name, path.stem),
    )


def load_client_contexts(context_dir: Path) -> list[ClientContext]:
    context_dir = resolve_existing_path(context_dir)
    if not context_dir.exists():
        return []

    contexts = []
    for path in sorted(context_dir.glob("*.json")):
        try:
            contexts.append(load_client_context(path))
        except Exception:
            continue
    return contexts


def decisive_token_overlap(left: set[str], right: set[str]) -> list[str]:
    overlap = sorted(left & right)
    if len(overlap) >= 2:
        return overlap
    if overlap and len(overlap[0]) >= 8:
        return overlap
    return []


def client_name_strong_match(notification_tokens: set[str], context: ClientContext) -> bool:
    if not notification_tokens or not context.name_tokens:
        return False
    return bool(decisive_token_overlap(notification_tokens, context.name_tokens))


def score_client_context(
    notification: dict[str, str],
    context: ClientContext,
    token_hints: dict[str, set[str]],
) -> ClientMatch:
    score = 0
    reasons: list[str] = []
    debug_parts: list[str] = []

    client_nif = notification_client_nif(notification)
    notif_tokens = notification_client_tokens(notification, token_hints=token_hints)
    filename_hits = decisive_token_overlap(notif_tokens, context.path_tokens)
    json_hits = sorted(notif_tokens & context.text_tokens)
    strong_name = client_name_strong_match(notif_tokens, context)
    nif_match = bool(client_nif and client_nif in context.normalized_nif_text)

    if nif_match:
        score += 80
        reasons.append("NIF en client_context")

    if filename_hits:
        score += 60
        reasons.append("tokens cliente en filename/json_path: " + ", ".join(filename_hits[:6]))

    if strong_name:
        score += 50
        reasons.append("nombre cliente fuerte")

    if json_hits:
        score += 25
        reasons.append("tokens cliente en JSON: " + ", ".join(json_hits[:6]))

    if nif_match and not filename_hits and not strong_name:
        score -= 40
        reasons.append("penalizacion NIF sin tokens de nombre")

    if client_nif and not nif_match and score >= 80:
        score = 79
        reasons.append("capado: contexto sin NIF de la notificacion")

    debug_parts.append(f"client_nif={client_nif or '-'}")
    debug_parts.append("notif_client_tokens=" + (",".join(sorted(notif_tokens)) or "-"))
    debug_parts.append("context_name_tokens=" + (",".join(sorted(context.name_tokens)) or "-"))
    debug_parts.append("filename_hits=" + (",".join(filename_hits) or "-"))
    debug_parts.append("json_hits=" + (",".join(json_hits[:12]) or "-"))

    return ClientMatch(
        context=context,
        score=max(0, min(score, 100)),
        reasons=reasons,
        debug=" | ".join(debug_parts),
    )


def best_client_context(
    notification: dict[str, str],
    contexts: list[ClientContext],
    token_hints: dict[str, set[str]],
) -> ClientMatch:
    best = ClientMatch(context=None, score=0, reasons=[], debug="")

    for context in contexts:
        current = score_client_context(notification, context, token_hints)
        if current.score > best.score:
            best = current

    return best


def index_cliente(row: dict[str, str] | None, columns: IndexColumns) -> str:
    if not row:
        return ""
    return first_present(row, columns.cliente)


def index_expediente(row: dict[str, str] | None, columns: IndexColumns) -> str:
    if not row:
        return ""
    return first_present(row, columns.expediente)


def index_path(row: dict[str, str] | None, columns: IndexColumns) -> str:
    if not row:
        return ""
    return first_present(row, columns.ruta)


def index_nifs(row: dict[str, str], columns: IndexColumns) -> set[str]:
    values = values_from_columns(row, columns.nif)
    return {nif for value in values for nif in [normalize_nif(value)] if nif}


def client_row_score(context: ClientContext, row: dict[str, str], columns: IndexColumns) -> tuple[int, str]:
    row_client_values = values_from_columns(row, columns.cliente)
    row_path_values = values_from_columns(row, columns.ruta)
    row_tokens = significant_tokens(unique_join(row_client_values), unique_join(row_path_values))
    context_tokens = context.name_tokens | context.path_tokens
    overlap = sorted(context_tokens & row_tokens)

    if not overlap:
        return 0, ""
    if len(overlap) >= 2 or any(len(token) >= 8 for token in overlap):
        return 50, "cliente en expediente vivo: " + ", ".join(overlap[:6])
    return 0, ""


def choose_expediente(
    notification: dict[str, str],
    client_match: ClientMatch,
    index_rows: list[dict[str, str]],
    columns: IndexColumns,
) -> Match:
    if not client_match.context:
        return Match(row=None, score=0, reasons=[], debug="")

    best = Match(row=None, score=0, reasons=[], debug="")
    notif_tokens = relevant_tokens(
        notification_value(notification, "title"),
        first_present(notification, ("procedure", "procedimiento", "numero_procedimiento")),
        first_present(notification, ("sender", "remitente", "organismo")),
    )

    notificaciones_fallback: Match | None = None

    for row in index_rows:
        base_score, base_reason = client_row_score(client_match.context, row, columns)
        if not base_score:
            continue

        expediente_text = unique_join(values_from_columns(row, columns.expediente))
        ruta_text = unique_join(values_from_columns(row, columns.ruta))
        row_tokens = relevant_tokens(expediente_text, ruta_text)
        token_hits = sorted(notif_tokens & row_tokens)
        score = base_score
        reasons = [base_reason]

        if token_hits:
            score += 20
            reasons.append("tokens expediente/ruta: " + ", ".join(token_hits[:6]))

        if "notificaciones" in normalize(expediente_text + " " + ruta_text).split():
            fallback = Match(row=row, score=max(score, 50), reasons=[base_reason, "fallback Notificaciones"])
            if notificaciones_fallback is None or fallback.score > notificaciones_fallback.score:
                notificaciones_fallback = fallback

        current = Match(row=row, score=min(score, 100), reasons=reasons)
        if current.score > best.score:
            best = current

    if best.score >= 50:
        return best
    if notificaciones_fallback:
        return notificaciones_fallback
    return Match(row=None, score=0, reasons=[], debug="cliente identificado sin expediente vivo claro")


def build_output_rows(
    notifications: list[dict[str, str]],
    index_rows: list[dict[str, str]],
    contexts: list[ClientContext],
) -> list[dict[str, str]]:
    output_rows: list[dict[str, str]] = []
    columns = detect_index_columns(index_rows)
    columns_detected = columns.describe()
    token_hints = build_nif_token_hints(notifications)

    for notification in notifications:
        client_match = best_client_context(notification, contexts, token_hints)
        expediente_match = choose_expediente(notification, client_match, index_rows, columns)
        client_context = client_match.context
        has_client = client_context is not None and client_match.score >= 80
        has_expediente = expediente_match.row is not None and expediente_match.score >= 50

        if has_client and has_expediente:
            estado = "pendiente_confirmacion"
            accion = "revisar_y_confirmar_destino"
        elif has_client:
            estado = "cliente_identificado"
            accion = "revisar_manualmente"
        else:
            estado = "requiere_revision"
            accion = "revisar_manualmente"

        cliente_sugerido = client_context.client_name if has_client and client_context else ""
        expediente_row = expediente_match.row if has_client and has_expediente else None
        match_debug = " | ".join(
            part
            for part in (client_match.debug, expediente_match.debug)
            if part
        )

        output_rows.append(
            {
                "notification_id": notification_value(notification, "notification_id"),
                "source": notification_value(notification, "source"),
                "received_at": notification_value(notification, "received_at"),
                "entity_nif": notification_value(notification, "entity_nif"),
                "entity_name": notification_value(notification, "entity_name"),
                "title": notification_value(notification, "title"),
                "download_status": notification_value(notification, "download_status"),
                "main_pdf_path": notification_value(notification, "main_pdf_path"),
                "receipt_path": notification_value(notification, "receipt_path"),
                "client_context_json": str(client_context.json_path) if has_client and client_context else "",
                "client_match_score": str(client_match.score),
                "client_match_reason": " | ".join(client_match.reasons),
                "cliente_sugerido": cliente_sugerido,
                "expediente_sugerido": index_expediente(expediente_row, columns) if expediente_row else "",
                "expediente_path": index_path(expediente_row, columns) if expediente_row else "",
                "match_score": str(expediente_match.score if has_expediente else client_match.score),
                "match_reason": " | ".join(expediente_match.reasons),
                "estado_asignacion": estado,
                "accion_sugerida": accion,
                "observaciones": "" if estado == "pendiente_confirmacion" else "Revisar destino antes de mover nada",
                "index_columns_detected": columns_detected,
                "match_debug": match_debug,
            }
        )

    return output_rows


def write_csv(rows: list[dict[str, str]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with output_path.open("w", newline="", encoding="utf-8-sig") as f:
        f.write("sep=;\n")
        writer = csv.DictWriter(
            f,
            fieldnames=OUTPUT_FIELDS,
            delimiter=";",
            lineterminator="\n",
        )
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(rows: list[dict[str, str]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Asignaciones"
    sheet.append(OUTPUT_FIELDS)

    for row in rows:
        sheet.append([row.get(field, "") for field in OUTPUT_FIELDS])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    sheet.freeze_panes = "A2"
    last_column = get_column_letter(len(OUTPUT_FIELDS))
    last_row = max(sheet.max_row, 1)
    sheet.auto_filter.ref = f"A1:{last_column}{last_row}"

    for column_idx, field in enumerate(OUTPUT_FIELDS, start=1):
        values = [field, *(str(row.get(field, "")) for row in rows[:200])]
        width = min(max(len(value) for value in values) + 2, 60)
        sheet.column_dimensions[get_column_letter(column_idx)].width = width

    workbook.save(output_path)


def printable_delimiter(delimiter: str) -> str:
    return "\\t" if delimiter == "\t" else delimiter


def print_summary(
    rows: list[dict[str, str]],
    inbox_csv: CsvReadResult,
    live_csv: CsvReadResult,
    contexts: list[ClientContext],
    index_columns: IndexColumns,
) -> None:
    counts: dict[str, int] = {}
    for row in rows:
        estado = row["estado_asignacion"]
        counts[estado] = counts.get(estado, 0) + 1

    print("Columnas detectadas live_expedientes:", index_columns.describe())
    print("Delimiter inbox:", printable_delimiter(inbox_csv.delimiter))
    print("Delimiter live_expedientes:", printable_delimiter(live_csv.delimiter))
    print("Total filas notifica:", len(inbox_csv.rows))
    print("Total filas expedientes:", len(live_csv.rows))
    print("Total contexts:", len(contexts))
    print("Resumen por estado_asignacion:", counts)
    print("Top 8:")
    ordered = sorted(rows, key=lambda row: int(row.get("client_match_score") or 0), reverse=True)
    for row in ordered[:8]:
        print(
            " - ".join(
                [
                    row.get("notification_id", ""),
                    row.get("estado_asignacion", ""),
                    row.get("client_match_score", ""),
                    row.get("cliente_sugerido", ""),
                    row.get("expediente_sugerido", ""),
                    row.get("client_match_reason", ""),
                ]
            )
        )


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Propone asignacion cliente/expediente para la bandeja exportada de Notifica."
    )
    parser.add_argument("--notifica-inbox", default=DEFAULT_NOTIFICA_INBOX)
    parser.add_argument("--live-expedientes", default=DEFAULT_LIVE_EXPEDIENTES)
    parser.add_argument("--client-context-dir", default=DEFAULT_CLIENT_CONTEXT_DIR)
    parser.add_argument("--out-csv", default=DEFAULT_OUT_CSV)
    parser.add_argument("--out-xlsx", default=DEFAULT_OUT_XLSX)
    args = parser.parse_args()

    notifica_inbox = resolve_existing_path(Path(args.notifica_inbox))
    live_expedientes = resolve_existing_path(Path(args.live_expedientes))
    client_context_dir = resolve_existing_path(Path(args.client_context_dir))
    out_csv = Path(args.out_csv)
    out_xlsx = Path(args.out_xlsx)

    inbox_csv = read_csv_flexible(notifica_inbox)
    live_csv = read_csv_flexible(live_expedientes)
    contexts = load_client_contexts(client_context_dir)
    index_columns = detect_index_columns(live_csv.rows)
    output_rows = build_output_rows(inbox_csv.rows, live_csv.rows, contexts)

    write_csv(output_rows, out_csv)
    write_xlsx(output_rows, out_xlsx)
    print_summary(output_rows, inbox_csv, live_csv, contexts, index_columns)
    print("CSV:", out_csv)
    print("XLSX:", out_xlsx)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())



