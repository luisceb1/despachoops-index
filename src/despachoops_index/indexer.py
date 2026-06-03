from __future__ import annotations

import sqlite3
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from despachoops_index.config import (
    DOCUMENT_EXTENSIONS,
    IndexOptions,
    LONG_PATH_THRESHOLD,
    TEXT_EXTENSIONS,
    TEXT_FTS_MAX,
    TEXT_PREVIEW_MAX,
)
from despachoops_index.hashutil import file_sha256
from despachoops_index.walk import iter_scan_paths, skip_reason


@dataclass(frozen=True)
class IndexResult:
    db_path: Path
    scanned: int
    indexed: int
    skipped_ignored: int
    skipped_unchanged: int
    read_errors: int
    with_text: int
    fts_enabled: bool
    limit_reached: bool


def build_index(options: IndexOptions) -> IndexResult:
    root = options.root
    if not root.exists():
        raise FileNotFoundError(f"No existe la carpeta raíz: {root}")

    options.db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(options.db_path)
    try:
        fts_enabled = _init_schema(conn)
        scanned = indexed = skipped = unchanged = read_errors = with_text = 0
        limit_reached = False

        for path in iter_scan_paths(root, options.scan_filters):
            scanned += 1
            if options.limit > 0 and indexed >= options.limit:
                limit_reached = True
                break

            ignored, reason = skip_reason(path, options.scan_filters)
            if ignored:
                _insert_ignored(conn, path, reason)
                skipped += 1
                continue

            if options.incremental:
                existing = _existing_row(conn, path)
                if existing and _unchanged(existing, path, options):
                    unchanged += 1
                    continue

            row, preview, full_text, err = _file_row(path, options)
            if err:
                read_errors += 1
            if preview or full_text:
                with_text += 1

            file_id = _upsert_file(conn, row)
            if options.include_text and (preview or full_text or err):
                _upsert_text(conn, file_id, preview, full_text, err)
                _upsert_fts(conn, fts_enabled, file_id, row, preview, full_text)
            indexed += 1

        conn.commit()
        return IndexResult(
            options.db_path,
            scanned,
            indexed,
            skipped,
            unchanged,
            read_errors,
            with_text,
            fts_enabled,
            limit_reached,
        )
    finally:
        conn.close()


def _init_schema(conn: sqlite3.Connection) -> bool:
    conn.execute("PRAGMA foreign_keys = ON")
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS files (
            id INTEGER PRIMARY KEY,
            path TEXT UNIQUE NOT NULL,
            name TEXT,
            extension TEXT,
            size_bytes INTEGER,
            mtime_iso TEXT,
            mtime_ns INTEGER,
            path_length INTEGER,
            parent_folder TEXT,
            is_temp_junk INTEGER DEFAULT 0,
            is_pdf INTEGER DEFAULT 0,
            is_docx INTEGER DEFAULT 0,
            is_xlsx INTEGER DEFAULT 0,
            is_txt INTEGER DEFAULT 0,
            is_csv INTEGER DEFAULT 0,
            read_error TEXT,
            indexed_at TEXT
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS file_text (
            file_id INTEGER PRIMARY KEY,
            text_preview TEXT,
            text_full TEXT,
            FOREIGN KEY(file_id) REFERENCES files(id)
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS ignored_files (
            id INTEGER PRIMARY KEY,
            path TEXT UNIQUE NOT NULL,
            name TEXT,
            reason TEXT,
            indexed_at TEXT
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS llm_enrichment (
            file_id INTEGER PRIMARY KEY,
            tipo_documental TEXT,
            area_sugerida TEXT,
            resumen TEXT,
            palabras_clave TEXT,
            confianza TEXT,
            necesita_revision INTEGER,
            modelo TEXT,
            status TEXT,
            error_message TEXT,
            enriched_at TEXT,
            FOREIGN KEY(file_id) REFERENCES files(id)
        )
        """
    )
    try:
        conn.execute(
            """
            CREATE VIRTUAL TABLE IF NOT EXISTS files_fts
            USING fts5(path, name, extension, text_preview, text_full)
            """
        )
        return True
    except sqlite3.OperationalError:
        return False


def _existing_row(conn: sqlite3.Connection, path: Path) -> sqlite3.Row | None:
    conn.row_factory = sqlite3.Row
    return conn.execute(
        "SELECT id, size_bytes, mtime_ns FROM files WHERE path = ?",
        (str(path.resolve()),),
    ).fetchone()


def _unchanged(existing: sqlite3.Row, path: Path, options: IndexOptions) -> bool:
    try:
        stat = path.stat()
    except OSError:
        return False
    return existing["size_bytes"] == stat.st_size and existing["mtime_ns"] == stat.st_mtime_ns


def _file_row(path: Path, options: IndexOptions) -> tuple[dict, str, str, str]:
    err = ""
    preview = ""
    full_text = ""
    try:
        stat = path.stat()
        if options.skip_large_files_mb > 0 and stat.st_size > options.skip_large_files_mb * 1024 * 1024:
            return _row_skip_large(path, stat), "", "", "omitido_tamano"
        ext = path.suffix.lower()
        rendered_path = str(path.resolve())
        row = {
            "path": rendered_path,
            "name": path.name,
            "extension": ext,
            "size_bytes": stat.st_size,
            "mtime_iso": datetime.fromtimestamp(stat.st_mtime).isoformat(timespec="seconds"),
            "mtime_ns": stat.st_mtime_ns,
            "path_length": len(rendered_path),
            "parent_folder": path.parent.name,
            "is_temp_junk": 0,
            "is_pdf": 1 if ext == ".pdf" else 0,
            "is_docx": 1 if ext == ".docx" else 0,
            "is_xlsx": 1 if ext in {".xlsx", ".xls"} else 0,
            "is_txt": 1 if ext == ".txt" else 0,
            "is_csv": 1 if ext == ".csv" else 0,
            "read_error": "",
            "indexed_at": datetime.now().isoformat(timespec="seconds"),
        }
        if options.include_text and ext in DOCUMENT_EXTENSIONS:
            preview, full_text, err = _extract_text(path, ext, options)
            row["read_error"] = err
        return row, preview, full_text, err
    except OSError as exc:
        return (
            {
                "path": str(path),
                "name": path.name,
                "extension": path.suffix.lower(),
                "size_bytes": 0,
                "mtime_iso": "",
                "mtime_ns": 0,
                "path_length": len(str(path)),
                "parent_folder": path.parent.name,
                "is_temp_junk": 0,
                "is_pdf": 0,
                "is_docx": 0,
                "is_xlsx": 0,
                "is_txt": 0,
                "is_csv": 0,
                "read_error": f"stat_error:{exc}",
                "indexed_at": datetime.now().isoformat(timespec="seconds"),
            },
            "",
            "",
            f"stat_error:{exc}",
        )


def _row_skip_large(path: Path, stat) -> dict:
    rendered = str(path.resolve())
    ext = path.suffix.lower()
    return {
        "path": rendered,
        "name": path.name,
        "extension": ext,
        "size_bytes": stat.st_size,
        "mtime_iso": datetime.fromtimestamp(stat.st_mtime).isoformat(timespec="seconds"),
        "mtime_ns": stat.st_mtime_ns,
        "path_length": len(rendered),
        "parent_folder": path.parent.name,
        "is_temp_junk": 0,
        "is_pdf": 1 if ext == ".pdf" else 0,
        "is_docx": 1 if ext == ".docx" else 0,
        "is_xlsx": 1 if ext in {".xlsx", ".xls"} else 0,
        "is_txt": 1 if ext == ".txt" else 0,
        "is_csv": 1 if ext == ".csv" else 0,
        "read_error": "omitido_tamano",
        "indexed_at": datetime.now().isoformat(timespec="seconds"),
    }


def _extract_text(path: Path, ext: str, options: IndexOptions) -> tuple[str, str, str]:
    try:
        if ext in TEXT_EXTENSIONS:
            text = path.read_text(encoding="utf-8", errors="ignore")
            return _clip(text)
        if ext == ".pdf":
            clipped = _clip(_read_pdf(path))
            if clipped[0] or not options.use_ocr_cache:
                return clipped
            return _clip(_ocr_cache_text(path, options))
        if ext == ".docx":
            return _clip(_read_docx(path))
        if ext in {".xlsx", ".xls"}:
            return _clip(_read_xlsx(path))
        if ext in {".png", ".jpg", ".jpeg", ".tif", ".tiff"} and options.use_ocr_cache:
            return _clip(_ocr_cache_text(path, options))
    except Exception as exc:  # noqa: BLE001
        return "", "", f"read_error:{exc}"
    return "", "", ""


def _ocr_cache_text(path: Path, options: IndexOptions) -> str:
    if not options.ocr_cache_dir:
        return ""
    cache = options.ocr_cache_dir / f"{file_sha256(path)}.txt"
    if cache.is_file():
        return cache.read_text(encoding="utf-8", errors="ignore")
    return ""


def _clip(text: str) -> tuple[str, str, str]:
    text = text.strip()
    if not text:
        return "", "", "sin_texto"
    preview = text[:TEXT_PREVIEW_MAX]
    full = text[:TEXT_FTS_MAX]
    return preview, full, ""


def _read_pdf(path: Path) -> str:
    from pypdf import PdfReader  # type: ignore

    reader = PdfReader(str(path))
    if reader.is_encrypted:
        return ""
    return "\n".join((page.extract_text() or "") for page in reader.pages)


def _read_docx(path: Path) -> str:
    from docx import Document  # type: ignore

    doc = Document(str(path))
    return "\n".join(p.text for p in doc.paragraphs if p.text)


def _read_xlsx(path: Path) -> str:
    from openpyxl import load_workbook  # type: ignore

    wb = load_workbook(str(path), read_only=True, data_only=True)
    parts: list[str] = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(max_row=150, values_only=True):
            for cell in row:
                if cell is not None:
                    parts.append(str(cell))
    wb.close()
    return " ".join(parts)


def _upsert_file(conn: sqlite3.Connection, row: dict) -> int:
    conn.execute(
        """
        INSERT INTO files (
            path, name, extension, size_bytes, mtime_iso, mtime_ns,
            path_length, parent_folder, is_temp_junk,
            is_pdf, is_docx, is_xlsx, is_txt, is_csv,
            read_error, indexed_at
        ) VALUES (
            :path, :name, :extension, :size_bytes, :mtime_iso, :mtime_ns,
            :path_length, :parent_folder, :is_temp_junk,
            :is_pdf, :is_docx, :is_xlsx, :is_txt, :is_csv,
            :read_error, :indexed_at
        )
        ON CONFLICT(path) DO UPDATE SET
            name=excluded.name, extension=excluded.extension,
            size_bytes=excluded.size_bytes, mtime_iso=excluded.mtime_iso,
            mtime_ns=excluded.mtime_ns, path_length=excluded.path_length,
            parent_folder=excluded.parent_folder,
            is_pdf=excluded.is_pdf, is_docx=excluded.is_docx,
            is_xlsx=excluded.is_xlsx, is_txt=excluded.is_txt,
            is_csv=excluded.is_csv,
            read_error=excluded.read_error, indexed_at=excluded.indexed_at
        """,
        row,
    )
    cur = conn.execute("SELECT id FROM files WHERE path = ?", (row["path"],))
    return int(cur.fetchone()[0])


def _upsert_text(conn: sqlite3.Connection, file_id: int, preview: str, full: str, err: str) -> None:
    conn.execute("DELETE FROM file_text WHERE file_id = ?", (file_id,))
    conn.execute(
        "INSERT INTO file_text (file_id, text_preview, text_full) VALUES (?, ?, ?)",
        (file_id, preview or err, full),
    )


def _upsert_fts(
    conn: sqlite3.Connection,
    enabled: bool,
    file_id: int,
    row: dict,
    preview: str,
    full: str,
) -> None:
    if not enabled:
        return
    conn.execute("DELETE FROM files_fts WHERE rowid = ?", (file_id,))
    conn.execute(
        """
        INSERT INTO files_fts (rowid, path, name, extension, text_preview, text_full)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (file_id, row["path"], row["name"], row["extension"], preview, full),
    )


def _insert_ignored(conn: sqlite3.Connection, path: Path, reason: str) -> None:
    now = datetime.now().isoformat(timespec="seconds")
    conn.execute(
        """
        INSERT OR IGNORE INTO ignored_files (path, name, reason, indexed_at)
        VALUES (?, ?, ?, ?)
        """,
        (str(path.resolve()), path.name, reason, now),
    )


def is_long_path(path_length: int) -> bool:
    return path_length >= LONG_PATH_THRESHOLD
