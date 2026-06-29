from __future__ import annotations

import csv
import sqlite3
from collections import Counter
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from despachoops_index.config import DOCUMENT_EXTENSIONS, LONG_PATH_THRESHOLD


@dataclass(frozen=True)
class DashboardResult:
    output_path: Path
    sheets: tuple[str, ...]


def build_dashboard(db_path: Path, output_path: Path) -> DashboardResult:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row

    try:
        summary = _summary(conn, db_path)
        by_ext = _by_extension(conn)
        long_paths = _long_paths(conn)
        sin_texto = _sin_texto(conn)
        pdfs = _pdfs(conn)
        duplicates = _duplicates(conn)
        ignored = _ignored(conn)

        ocr_cola = _ocr_cola(db_path)
        ocr_documentos = _ocr_documentos(conn)
        ocr_tipos = _ocr_tipos(conn)
        ocr_extracciones = _ocr_extracciones(conn)
        ocr_plazos = _ocr_plazos(conn)

        wb = Workbook()
        wb.remove(wb.active)

        _write_sheet(wb, "Resumen", ["metrica", "valor"], summary)
        _write_sheet(wb, "Extensiones", ["extension", "archivos"], by_ext)
        _write_sheet(wb, "Rutas_Largas", ["path", "path_length", "name"], long_paths)
        _write_sheet(
            wb,
            "Sin_Texto",
            ["path", "extension", "read_error", "size_bytes", "name"],
            sin_texto,
        )
        _write_sheet(
            wb,
            "PDFs",
            ["path", "size_bytes", "mtime_iso", "read_error"],
            pdfs,
        )
        _write_sheet(
            wb,
            "Duplicados_Probables",
            ["name", "size_bytes", "copias", "paths"],
            duplicates,
        )
        _write_sheet(
            wb,
            "Archivos_Temporales_Ignorados",
            ["path", "name", "reason"],
            ignored,
        )

        _write_sheet(
            wb,
            "OCR_Cola",
            ["estado", "extension", "archivos"],
            ocr_cola,
        )
        _write_sheet(
            wb,
            "OCR_Documentos",
            [
                "tipo_documento_detectado",
                "posible_cliente",
                "posible_expediente",
                "archivo_original",
                "ruta_original",
                "estado_ocr",
                "status_ocr",
                "texto_chars",
                "confianza_aproximada",
                "num_paginas",
            ],
            ocr_documentos,
        )
        _write_sheet(
            wb,
            "OCR_Tipos",
            ["tipo_documento_detectado", "archivos"],
            ocr_tipos,
        )
        _write_sheet(
            wb,
            "OCR_Extracciones",
            ["metrica", "valor"],
            ocr_extracciones,
        )
        _write_sheet(
            wb,
            "OCR_Plazos",
            [
                "tipo_documento_detectado",
                "posible_cliente",
                "archivo_original",
                "ruta_original",
                "plazos_detectados_json",
                "fechas_detectadas_json",
                "procedimientos_detectados_json",
            ],
            ocr_plazos,
        )

        wb.save(output_path)

        return DashboardResult(
            output_path,
            (
                "Resumen",
                "Extensiones",
                "Rutas_Largas",
                "Sin_Texto",
                "PDFs",
                "Duplicados_Probables",
                "Archivos_Temporales_Ignorados",
                "OCR_Cola",
                "OCR_Documentos",
                "OCR_Tipos",
                "OCR_Extracciones",
                "OCR_Plazos",
            ),
        )

    finally:
        conn.close()


def _summary(conn: sqlite3.Connection, db_path: Path) -> list[dict[str, str]]:
    total = conn.execute("SELECT COUNT(*) FROM files").fetchone()[0]

    with_text = conn.execute(
        """
        SELECT COUNT(*) FROM files f
        JOIN file_text t ON t.file_id = f.id
        WHERE COALESCE(t.text_preview, '') != ''
          AND COALESCE(f.read_error, '') = ''
        """
    ).fetchone()[0]

    errors = conn.execute(
        "SELECT COUNT(*) FROM files WHERE read_error IS NOT NULL AND read_error != ''"
    ).fetchone()[0]

    long_count = conn.execute(
        "SELECT COUNT(*) FROM files WHERE path_length >= ?",
        (LONG_PATH_THRESHOLD,),
    ).fetchone()[0]

    ignored = conn.execute("SELECT COUNT(*) FROM ignored_files").fetchone()[0]

    dup_groups = conn.execute(
        """
        SELECT COUNT(*) FROM (
            SELECT name, size_bytes
            FROM files
            GROUP BY name, size_bytes
            HAVING COUNT(*) > 1
        )
        """
    ).fetchone()[0]

    ocr_metrics = _ocr_summary_metrics(conn, db_path)

    return [
        {"metrica": "total_indexados", "valor": str(total)},
        {"metrica": "con_texto", "valor": str(with_text)},
        {"metrica": "read_error", "valor": str(errors)},
        {"metrica": "rutas_largas", "valor": str(long_count)},
        {"metrica": "grupos_duplicados_probables", "valor": str(dup_groups)},
        {"metrica": "archivos_ignorados", "valor": str(ignored)},
        {"metrica": "umbral_ruta_larga", "valor": str(LONG_PATH_THRESHOLD)},
        {"metrica": "ocr_jobs_total", "valor": str(ocr_metrics.get("ocr_jobs_total", 0))},
        {"metrica": "ocr_ok", "valor": str(ocr_metrics.get("ocr_ok", 0))},
        {"metrica": "ocr_cache", "valor": str(ocr_metrics.get("ocr_cache", 0))},
        {"metrica": "ocr_pendiente", "valor": str(ocr_metrics.get("ocr_pendiente", 0))},
        {"metrica": "ocr_sin_texto", "valor": str(ocr_metrics.get("ocr_sin_texto", 0))},
        {"metrica": "ocr_error", "valor": str(ocr_metrics.get("ocr_error", 0))},
        {"metrica": "ocr_no_encontrado", "valor": str(ocr_metrics.get("ocr_no_encontrado", 0))},
        {"metrica": "ocr_documents_total", "valor": str(ocr_metrics.get("ocr_documents_total", 0))},
        {"metrica": "ocr_documents_con_texto", "valor": str(ocr_metrics.get("ocr_documents_con_texto", 0))},
        {"metrica": "ocr_con_nif_cif", "valor": str(ocr_metrics.get("ocr_con_nif_cif", 0))},
        {"metrica": "ocr_con_importes", "valor": str(ocr_metrics.get("ocr_con_importes", 0))},
        {"metrica": "ocr_con_plazos", "valor": str(ocr_metrics.get("ocr_con_plazos", 0))},
    ]


def _by_extension(conn: sqlite3.Connection) -> list[dict[str, str]]:
    rows = conn.execute(
        """
        SELECT COALESCE(NULLIF(extension, ''), '(sin)') AS ext, COUNT(*) AS c
        FROM files
        GROUP BY ext
        ORDER BY c DESC
        """
    ).fetchall()

    return [{"extension": r[0], "archivos": str(r[1])} for r in rows]


def _long_paths(conn: sqlite3.Connection) -> list[dict[str, str]]:
    rows = conn.execute(
        """
        SELECT path, path_length, name
        FROM files
        WHERE path_length >= ?
        ORDER BY path_length DESC
        LIMIT 5000
        """,
        (LONG_PATH_THRESHOLD,),
    ).fetchall()

    return [
        {
            "path": r[0],
            "path_length": str(r[1]),
            "name": r[2],
        }
        for r in rows
    ]


def _sin_texto(conn: sqlite3.Connection) -> list[dict[str, str]]:
    placeholders = ",".join("?" for _ in DOCUMENT_EXTENSIONS)

    rows = conn.execute(
        f"""
        SELECT
            f.path,
            f.extension,
            COALESCE(f.read_error, '') AS read_error,
            f.size_bytes,
            f.name,
            COALESCE(t.text_preview, '') AS text_preview
        FROM files f
        LEFT JOIN file_text t ON t.file_id = f.id
        WHERE LOWER(f.extension) IN ({placeholders})
          AND (
                COALESCE(f.read_error, '') != ''
             OR t.file_id IS NULL
             OR TRIM(COALESCE(t.text_preview, '')) = ''
             OR LOWER(TRIM(t.text_preview)) = 'sin_texto'
             OR LOWER(TRIM(t.text_preview)) LIKE 'read_error:%'
          )
        ORDER BY f.path
        LIMIT 5000
        """,
        tuple(ext.lower() for ext in DOCUMENT_EXTENSIONS),
    ).fetchall()

    return [
        {
            "path": r[0],
            "extension": r[1],
            "read_error": r[2] or _preview_as_error(r[5]),
            "size_bytes": str(r[3] if r[3] is not None else ""),
            "name": r[4] or "",
        }
        for r in rows
    ]


def _preview_as_error(preview: str) -> str:
    p = (preview or "").strip().lower()
    if p in {"sin_texto", "read_error"} or p.startswith("read_error:"):
        return preview.strip()
    return ""


def _pdfs(conn: sqlite3.Connection) -> list[dict[str, str]]:
    rows = conn.execute(
        """
        SELECT path, size_bytes, mtime_iso, COALESCE(read_error, '') AS read_error
        FROM files
        WHERE extension = '.pdf'
        ORDER BY path
        LIMIT 10000
        """
    ).fetchall()

    return [
        {
            "path": r[0],
            "size_bytes": str(r[1]),
            "mtime_iso": r[2],
            "read_error": r[3],
        }
        for r in rows
    ]


def _duplicates(conn: sqlite3.Connection) -> list[dict[str, str]]:
    groups = conn.execute(
        """
        SELECT name, size_bytes, COUNT(*) AS c
        FROM files
        GROUP BY name, size_bytes
        HAVING c > 1
        ORDER BY c DESC, name
        LIMIT 2000
        """
    ).fetchall()

    result: list[dict[str, str]] = []

    for name, size, count in groups:
        paths = conn.execute(
            """
            SELECT path
            FROM files
            WHERE name = ? AND size_bytes = ?
            ORDER BY path
            """,
            (name, size),
        ).fetchall()

        result.append(
            {
                "name": name,
                "size_bytes": str(size),
                "copias": str(count),
                "paths": " | ".join(p[0] for p in paths),
            }
        )

    return result


def _ignored(conn: sqlite3.Connection) -> list[dict[str, str]]:
    try:
        rows = conn.execute(
            """
            SELECT path, name, reason
            FROM ignored_files
            ORDER BY path
            LIMIT 5000
            """
        ).fetchall()
    except sqlite3.OperationalError:
        return []

    return [
        {
            "path": r[0],
            "name": r[1],
            "reason": r[2],
        }
        for r in rows
    ]


def _ocr_queue_path(db_path: Path) -> Path:
    return db_path.parent / "ocr_jobs.csv"


def _load_ocr_jobs(db_path: Path) -> list[dict[str, str]]:
    q = _ocr_queue_path(db_path)
    if not q.exists():
        return []

    with q.open("r", encoding="utf-8-sig", newline="") as fh:
        return list(csv.DictReader(fh))


def _table_exists(conn: sqlite3.Connection, table: str) -> bool:
    row = conn.execute(
        """
        SELECT name
        FROM sqlite_master
        WHERE type = 'table' AND name = ?
        """,
        (table,),
    ).fetchone()

    return row is not None


def _ocr_summary_metrics(conn: sqlite3.Connection, db_path: Path) -> dict[str, int]:
    rows = _load_ocr_jobs(db_path)

    metrics = {
        "ocr_jobs_total": len(rows),
        "ocr_ok": 0,
        "ocr_cache": 0,
        "ocr_pendiente": 0,
        "ocr_sin_texto": 0,
        "ocr_error": 0,
        "ocr_no_encontrado": 0,
        "ocr_documents_total": 0,
        "ocr_documents_con_texto": 0,
        "ocr_con_nif_cif": 0,
        "ocr_con_importes": 0,
        "ocr_con_plazos": 0,
    }

    for row in rows:
        estado = row.get("estado", "")

        if estado == "ocr_ok":
            metrics["ocr_ok"] += 1
        elif estado == "ocr_cache":
            metrics["ocr_cache"] += 1
        elif estado == "pendiente":
            metrics["ocr_pendiente"] += 1
        elif estado == "sin_texto":
            metrics["ocr_sin_texto"] += 1
        elif estado == "error":
            metrics["ocr_error"] += 1
        elif estado == "no_encontrado":
            metrics["ocr_no_encontrado"] += 1

    if _table_exists(conn, "ocr_documents"):
        metrics["ocr_documents_total"] = conn.execute(
            "SELECT COUNT(*) FROM ocr_documents"
        ).fetchone()[0]

        metrics["ocr_documents_con_texto"] = conn.execute(
            "SELECT COUNT(*) FROM ocr_documents WHERE texto_chars > 0"
        ).fetchone()[0]

        metrics["ocr_con_nif_cif"] = conn.execute(
            "SELECT COUNT(*) FROM ocr_documents WHERE nifs_cifs_detectados_json <> '[]'"
        ).fetchone()[0]

        metrics["ocr_con_importes"] = conn.execute(
            "SELECT COUNT(*) FROM ocr_documents WHERE importes_detectados_json <> '[]'"
        ).fetchone()[0]

        metrics["ocr_con_plazos"] = conn.execute(
            "SELECT COUNT(*) FROM ocr_documents WHERE plazos_detectados_json <> '[]'"
        ).fetchone()[0]

    return metrics


def _ocr_cola(db_path: Path) -> list[dict[str, str]]:
    rows = _load_ocr_jobs(db_path)
    counts: Counter[tuple[str, str]] = Counter()

    for row in rows:
        estado = row.get("estado", "") or "(sin_estado)"
        ext = row.get("extension", "") or "(sin_ext)"
        counts[(estado, ext)] += 1

    return [
        {
            "estado": estado,
            "extension": ext,
            "archivos": str(count),
        }
        for (estado, ext), count in sorted(
            counts.items(),
            key=lambda x: (-x[1], x[0][0], x[0][1]),
        )
    ]


def _ocr_documentos(conn: sqlite3.Connection) -> list[dict[str, str]]:
    if not _table_exists(conn, "ocr_documents"):
        return []

    rows = conn.execute(
        """
        SELECT
            tipo_documento_detectado,
            posible_cliente,
            posible_expediente,
            archivo_original,
            ruta_original,
            estado_ocr,
            status_ocr,
            texto_chars,
            confianza_aproximada,
            num_paginas
        FROM ocr_documents
        ORDER BY updated_at DESC
        LIMIT 10000
        """
    ).fetchall()

    return [
        {
            "tipo_documento_detectado": r[0] or "",
            "posible_cliente": r[1] or "",
            "posible_expediente": r[2] or "",
            "archivo_original": r[3] or "",
            "ruta_original": r[4] or "",
            "estado_ocr": r[5] or "",
            "status_ocr": r[6] or "",
            "texto_chars": str(r[7] or 0),
            "confianza_aproximada": str(r[8] if r[8] is not None else ""),
            "num_paginas": str(r[9] if r[9] is not None else ""),
        }
        for r in rows
    ]


def _ocr_tipos(conn: sqlite3.Connection) -> list[dict[str, str]]:
    if not _table_exists(conn, "ocr_documents"):
        return []

    rows = conn.execute(
        """
        SELECT tipo_documento_detectado, COUNT(*) AS c
        FROM ocr_documents
        GROUP BY tipo_documento_detectado
        ORDER BY c DESC
        LIMIT 100
        """
    ).fetchall()

    return [
        {
            "tipo_documento_detectado": r[0] or "(sin_tipo)",
            "archivos": str(r[1]),
        }
        for r in rows
    ]


def _ocr_extracciones(conn: sqlite3.Connection) -> list[dict[str, str]]:
    if not _table_exists(conn, "ocr_documents"):
        return []

    metrics = [
        ("ocr_documents_total", "SELECT COUNT(*) FROM ocr_documents"),
        ("con_texto", "SELECT COUNT(*) FROM ocr_documents WHERE texto_chars > 0"),
        (
            "con_nif_cif",
            "SELECT COUNT(*) FROM ocr_documents WHERE nifs_cifs_detectados_json <> '[]'",
        ),
        (
            "con_importes",
            "SELECT COUNT(*) FROM ocr_documents WHERE importes_detectados_json <> '[]'",
        ),
        (
            "con_plazos",
            "SELECT COUNT(*) FROM ocr_documents WHERE plazos_detectados_json <> '[]'",
        ),
        (
            "con_fechas",
            "SELECT COUNT(*) FROM ocr_documents WHERE fechas_detectadas_json <> '[]'",
        ),
        (
            "con_procedimientos",
            "SELECT COUNT(*) FROM ocr_documents WHERE procedimientos_detectados_json <> '[]'",
        ),
    ]

    out = []

    for name, sql in metrics:
        value = conn.execute(sql).fetchone()[0]
        out.append(
            {
                "metrica": name,
                "valor": str(value),
            }
        )

    return out


def _ocr_plazos(conn: sqlite3.Connection) -> list[dict[str, str]]:
    if not _table_exists(conn, "ocr_documents"):
        return []

    rows = conn.execute(
        """
        SELECT
            tipo_documento_detectado,
            posible_cliente,
            archivo_original,
            ruta_original,
            plazos_detectados_json,
            fechas_detectadas_json,
            procedimientos_detectados_json
        FROM ocr_documents
        WHERE plazos_detectados_json <> '[]'
        ORDER BY updated_at DESC
        LIMIT 5000
        """
    ).fetchall()

    return [
        {
            "tipo_documento_detectado": r[0] or "",
            "posible_cliente": r[1] or "",
            "archivo_original": r[2] or "",
            "ruta_original": r[3] or "",
            "plazos_detectados_json": r[4] or "[]",
            "fechas_detectadas_json": r[5] or "[]",
            "procedimientos_detectados_json": r[6] or "[]",
        }
        for r in rows
    ]


def _write_sheet(
    wb: Workbook,
    title: str,
    headers: list[str],
    rows: list[dict[str, str]],
) -> None:
    ws = wb.create_sheet(title)
    ws.append(headers)

    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    for idx, header in enumerate(headers, start=1):
        values = [header] + [str(row.get(header, "")) for row in rows]
        width = min(60, max(10, max(len(v) for v in values) + 2))
        ws.column_dimensions[get_column_letter(idx)].width = width