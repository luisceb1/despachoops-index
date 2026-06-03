from datetime import datetime
from pathlib import Path

import yaml
from openpyxl import load_workbook

from despachoops_index.cli import main
from despachoops_index.config import IndexOptions, load_app_config
from despachoops_index.indexer import build_index


def _write_cfg(
    tmp_path: Path,
    *,
    reports: Path | None = None,
    latest: Path | None = None,
) -> Path:
    data = tmp_path / "data"
    payload = {
        "scan_root": str(tmp_path / "Clientes"),
        "data_dir": str(data),
    }
    if reports is not None:
        payload["reports_dir"] = str(reports)
    if latest is not None:
        payload["latest_dir"] = str(latest)
    cfg_path = tmp_path / "config.yaml"
    cfg_path.write_text(yaml.dump(payload), encoding="utf-8")
    return cfg_path


def _seed_db(root: Path, db: Path) -> None:
    root.mkdir(parents=True, exist_ok=True)
    (root / "doc.txt").write_text("contenido", encoding="utf-8")
    build_index(IndexOptions(root=root, db_path=db, include_text=True))


def test_dashboard_without_output_creates_timestamped_xlsx(tmp_path: Path, monkeypatch):
    cfg_path = _write_cfg(tmp_path, reports=tmp_path / "reports", latest=tmp_path / "latest")
    app = load_app_config(cfg_path)
    scan = tmp_path / "Clientes"
    scan.mkdir()
    db = app.index_db_path
    _seed_db(scan, db)

    fixed = datetime(2026, 6, 3, 19, 35, 0)
    monkeypatch.setattr(
        "despachoops_index.config.datetime",
        type("DT", (), {"now": staticmethod(lambda: fixed)}),
    )

    assert main(["--config", str(cfg_path), "dashboard"]) == 0
    expected = tmp_path / "reports" / "index_dashboard_20260603_193500.xlsx"
    assert expected.exists()
    wb = load_workbook(expected, read_only=True)
    assert "Resumen" in wb.sheetnames
    wb.close()


def test_dashboard_publish_latest(tmp_path: Path):
    index = tmp_path / "Index"
    cfg_path = _write_cfg(
        tmp_path,
        reports=index / "reports",
        latest=index / "latest",
    )
    app = load_app_config(cfg_path)
    scan = tmp_path / "Clientes"
    scan.mkdir()
    db = app.index_db_path
    _seed_db(scan, db)

    reports_out = index / "reports" / "index_dashboard_20000.xlsx"
    latest_out = index / "latest" / "index_dashboard.xlsx"

    assert main(
        [
            "--config",
            str(cfg_path),
            "dashboard",
            "--output",
            str(reports_out),
            "--publish-latest",
        ]
    ) == 0
    assert reports_out.exists()
    assert latest_out.exists()
    assert latest_out.read_bytes() == reports_out.read_bytes()


def test_dashboard_with_explicit_output_in_latest(tmp_path: Path):
    index = tmp_path / "Index"
    cfg_path = _write_cfg(
        tmp_path,
        reports=index / "reports",
        latest=index / "latest",
    )
    app = load_app_config(cfg_path)
    scan = tmp_path / "Clientes"
    scan.mkdir()
    db = app.index_db_path
    _seed_db(scan, db)

    explicit = index / "latest" / "index_dashboard.xlsx"
    assert main(["--config", str(cfg_path), "dashboard", "--output", str(explicit)]) == 0
    assert explicit.exists()
