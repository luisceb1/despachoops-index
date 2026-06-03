from pathlib import Path

import pytest
from openpyxl import load_workbook

from despachoops_index.config import IndexOptions, LONG_PATH_THRESHOLD
from despachoops_index.dashboard import build_dashboard
from despachoops_index.indexer import build_index

EXPECTED_SHEETS = (
    "Resumen",
    "Extensiones",
    "Rutas_Largas",
    "Sin_Texto",
    "PDFs",
    "Duplicados_Probables",
    "Archivos_Temporales_Ignorados",
)


@pytest.fixture
def indexed_db(tmp_path: Path) -> tuple[Path, Path]:
    root = tmp_path / "docs"
    root.mkdir()
    payload = b"%PDF-1.4 minimal"
    (root / "mismo.pdf").write_bytes(payload)
    dup_dir = root / "cliente"
    dup_dir.mkdir()
    (dup_dir / "mismo.pdf").write_bytes(payload)
    long_dir = root / "expediente"
    long_dir.mkdir()
    (long_dir / ("a" * 200 + ".txt")).write_text("x", encoding="utf-8")
    (root / "~$lock.doc").write_bytes(b"1")
    (root / "contrato.txt").write_text("texto", encoding="utf-8")

    db = tmp_path / "index.sqlite"
    build_index(IndexOptions(root=root, db_path=db, include_text=True))
    return db, tmp_path / "reports" / "dash.xlsx"


def test_dashboard_all_sheets(indexed_db):
    db, out = indexed_db
    result = build_dashboard(db, out)
    assert out.exists()
    assert result.sheets == EXPECTED_SHEETS

    wb = load_workbook(out, read_only=True)
    assert set(wb.sheetnames) == set(EXPECTED_SHEETS)

    resumen = list(wb["Resumen"].iter_rows(min_row=2, values_only=True))
    metrics = {row[0]: row[1] for row in resumen if row[0]}
    assert int(metrics.get("total_indexados", 0)) >= 3

    dup_rows = list(wb["Duplicados_Probables"].iter_rows(min_row=2, values_only=True))
    assert any(row and row[0] == "mismo.pdf" for row in dup_rows)

    long_rows = list(wb["Rutas_Largas"].iter_rows(min_row=2, values_only=True))
    assert any(row and int(row[1]) >= LONG_PATH_THRESHOLD for row in long_rows if row[1])

    ignored = list(wb["Archivos_Temporales_Ignorados"].iter_rows(min_row=2, values_only=True))
    assert "~$lock.doc" not in {r[1] for r in ignored if r and len(r) > 1}

    pdf_rows = list(wb["PDFs"].iter_rows(min_row=2, values_only=True))
    assert len(pdf_rows) >= 2

    wb.close()
